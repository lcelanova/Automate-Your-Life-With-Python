import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
import string
from openpyxl.styles import PatternFill, Font


file = pd.read_excel('supermarket_sales.xlsx')

pivote_table = file.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)

pivote_table.to_excel('sales_2021.xlsx', startrow=4, sheet_name='Report')

workbook = load_workbook('sales_2021.xlsx')

tab = workbook['Report']

min_column = workbook.active.min_column
max_column = workbook.active.max_column
min_row = workbook.active.min_row
max_row = workbook.active.max_row

#plotting

barchart = BarChart()

data = Reference(tab, min_col=min_column+1, max_col=max_column, min_row=min_row, max_row=max_row)
categories = Reference(tab, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)
barchart.style = 2
barchart.title = 'Sales'

tab.add_chart(barchart, 'B12')

abc = list(string.ascii_uppercase)
abc_excel = abc[0:max_column]

for i in abc_excel:
	if i != 'A':
		tab[f'{i}{max_row + 1}'] = f'=SUM({i}{min_row + 1}:{i}{max_row})'
		tab[f'{i}{max_row + 1}'].style = 'Currency'

tab[f'{abc_excel[0]}{max_row + 1}'] = 'Total'

tab['A1'] = 'Report'
tab['A2'] = '2021'

tab['A1'].font = Font('Arial', bold=True, size=20)
tab['A2'].font = Font('Arial', bold=True, size=12)

workbook.save('sales_2021.xlsx')

